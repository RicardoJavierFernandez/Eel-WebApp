import eel
import openpyxl

eel.init('web')


web_app_options = {
    # 'mode': "chrome", #or "chrome-app"
    'host': '127.0.0.1', # 'localhost',
    'port': 8000
}

@eel.expose
def create_excel_file():
    wb = openpyxl.Workbook()
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    ws['A1'] = 'This is a test'
    ws['B1'] = '=100 * 23'

    wb.save("C:/Users/Ricardo/Downloads/EelTest.xlsx")

@eel.expose
def create_broker_file():
    wb = openpyxl.Workbook()
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    ws['A1'] = 'This file was made using eel'
    ws['A2'] = '=100*100'

    wb.save("C:/Users/Ricardo/Downloads/BrokerEmails.xlsx")

eel.start('index.html', options=web_app_options)


